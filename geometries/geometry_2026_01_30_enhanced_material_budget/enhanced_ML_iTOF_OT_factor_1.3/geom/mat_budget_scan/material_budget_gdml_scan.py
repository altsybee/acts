#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GDML material budget scan (X/X0) with ROOT TGeo (PyROOT).

What it does:
  - traces straight rays from an origin for a grid of (eta,phi)
  - accumulates X/X0 per material
  - stacked plot vs eta (default) or vs phi
  - saves PNG and ROOT (TH1D per material + total)
  - optional TH2D total map vs (eta,phi)
  - optional traversal tables (txt + xlsx) for selected rays (pairwise or cartesian)

Key fix for assemblies / physvol like FT3V_2:
  - ROOT may flatten GDML <assembly>, so physvol placement name may not show up in navigator path.
  - We parse GDML and expand requested physvol names into a set of descendant volume/assembly names.
  - During tracing we "collect" if current TGeoVolume name is in that expanded set.

Defaults:
  origin = (0, 30, 0) cm
  R in [0, 90] cm
  z in [-300, 300] cm
  collect physvols = A3IP_1, TRKV_2, IOTOFV_2, FT3V_2
"""

import math
import argparse
from collections import defaultdict, Counter
from fnmatch import fnmatch
import re
from typing import Dict, List, Tuple, Optional, Any, Set
import xml.etree.ElementTree as ET

import numpy as np
import matplotlib.pyplot as plt


# -----------------------------
# Kinematics helpers
# -----------------------------
def eta_phi_to_dir(eta: float, phi: float) -> Tuple[float, float, float]:
    theta = 2.0 * math.atan(math.exp(-eta))  # theta from +z
    st = math.sin(theta)
    ct = math.cos(theta)
    return (math.cos(phi) * st, math.sin(phi) * st, ct)


def linspace_including_end(a: float, b: float, n: int) -> np.ndarray:
    if n == 1:
        return np.array([0.5 * (a + b)], dtype=float)
    return np.linspace(a, b, n, dtype=float)


def parse_csv_floats(s: Optional[str]) -> List[float]:
    if not s:
        return []
    out: List[float] = []
    for tok in s.split(","):
        tok = tok.strip()
        if not tok:
            continue
        out.append(float(tok))
    return out


def parse_csv_strings(s: Optional[str]) -> List[str]:
    if not s:
        return []
    return [t.strip() for t in s.split(",") if t.strip()]


# -----------------------------
# ROOT geometry loading
# -----------------------------
def load_root_geometry(gdml_path: str):
    try:
        import ROOT  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "PyROOT (module 'ROOT') is required. Run in an environment with ROOT+GDML+PyROOT.\n"
            f"Import error: {e}"
        )

    geo = ROOT.TGeoManager.Import(gdml_path)
    if not geo:
        geo = getattr(ROOT, "gGeoManager", None)

    if not geo:
        raise RuntimeError("Failed to import GDML into ROOT.TGeoManager.")
    if not geo.GetTopVolume():
        raise RuntimeError("No top volume found after GDML import.")

    try:
        if not geo.IsClosed():
            geo.CloseGeometry()
    except Exception:
        pass

    return ROOT, geo


# -----------------------------
# GDML parsing to expand physvol -> descendant volume/assembly names
# -----------------------------
def _strip_ns(tag: str) -> str:
    # "{ns}volume" -> "volume"
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def parse_gdml_structure(gdml_path: str) -> Tuple[Dict[str, str], Dict[str, List[str]], Set[str]]:
    """
    Returns:
      physvol_to_ref:  physvol_name -> referenced structure name (volume/assembly) via <volumeref ref="...">
      container_to_children: container_name (volume or assembly) -> list of referenced child container names
      all_containers: set of all volume/assembly names defined in <structure>
    """
    tree = ET.parse(gdml_path)
    root = tree.getroot()

    # Find <structure>
    structure = None
    for ch in root:
        if _strip_ns(ch.tag) == "structure":
            structure = ch
            break
    if structure is None:
        return {}, {}, set()

    all_containers: Set[str] = set()
    container_to_children: Dict[str, List[str]] = defaultdict(list)

    # First pass: collect volumes/assemblies + their children refs
    for elem in structure:
        tag = _strip_ns(elem.tag)
        if tag not in ("volume", "assembly"):
            continue
        name = elem.attrib.get("name", "")
        if not name:
            continue
        all_containers.add(name)

        # children physvol inside this container
        for pv in elem:
            if _strip_ns(pv.tag) != "physvol":
                continue
            # child can be via <volumeref ref="...">
            ref = None
            for sub in pv:
                if _strip_ns(sub.tag) == "volumeref":
                    ref = sub.attrib.get("ref")
                    break
            if ref:
                container_to_children[name].append(ref)

    # Second pass: map physvol names -> volumeref ref
    physvol_to_ref: Dict[str, str] = {}
    for elem in structure:
        tag = _strip_ns(elem.tag)
        if tag not in ("volume", "assembly"):
            continue
        for pv in elem:
            if _strip_ns(pv.tag) != "physvol":
                continue
            pv_name = pv.attrib.get("name", "")
            if not pv_name:
                continue
            ref = None
            for sub in pv:
                if _strip_ns(sub.tag) == "volumeref":
                    ref = sub.attrib.get("ref")
                    break
            if ref:
                physvol_to_ref[pv_name] = ref

    return physvol_to_ref, dict(container_to_children), all_containers


def expand_descendants(start_ref: str, container_to_children: Dict[str, List[str]]) -> Set[str]:
    """
    DFS from a referenced container name (volume or assembly),
    returns set of all containers reachable including the start itself.
    """
    out: Set[str] = set()
    stack = [start_ref]
    while stack:
        cur = stack.pop()
        if cur in out:
            continue
        out.add(cur)
        for ch in container_to_children.get(cur, []):
            if ch not in out:
                stack.append(ch)
    return out


# -----------------------------
# Matching helpers (copies)
# -----------------------------
_TGEO_INSTANCE_RE = re.compile(r"([;#]\d+)$")  # ";1", "#12" at end


def norm_name(s: str) -> str:
    """
    Normalize ROOT/TGeo node/volume names:
      - strip ROOT instance suffixes ;N or #N (repeated)
      - strip extra copy-number tokens "_N" from the end, but keep ONE last numeric token
        (so physvol base names like FT3V_2 are preserved)
    """
    s = (s or "").strip()

    while True:
        m = _TGEO_INSTANCE_RE.search(s)
        if not m:
            break
        s = s[:m.start()]

    parts = s.split("_")

    def is_num(tok: str) -> bool:
        return tok.isdigit()

    while len(parts) >= 3 and is_num(parts[-1]) and is_num(parts[-2]):
        parts.pop()

    return "_".join(parts)


def is_collected_by_patterns(node_name: str, path: str, volume_name: str, patterns: List[str]) -> bool:
    """
    Pattern matching against:
      - node name (raw + normalized)
      - volume name (raw + normalized)
      - full path (raw)
    Plus: if pattern has no wildcards and no '/', allow substring-in-path.
    """
    node_raw = (node_name or "").strip()
    vol_raw = (volume_name or "").strip()
    node_n = norm_name(node_raw)
    vol_n = norm_name(vol_raw)
    path = (path or "").strip()

    for p in patterns:
        p = (p or "").strip()
        if not p:
            continue

        if fnmatch(node_raw, p) or fnmatch(node_n, p):
            return True
        if fnmatch(vol_raw, p) or fnmatch(vol_n, p):
            return True
        if path and fnmatch(path, p):
            return True

        has_wild = any(ch in p for ch in ["*", "?", "["])
        has_slash = "/" in p
        if (not has_wild) and (not has_slash) and path:
            if p in path:
                return True

    return False


def sanitize_root_name(s: str, prefix: str = "h_") -> str:
    s2 = re.sub(r"[^A-Za-z0-9_]+", "_", (s or "").strip())
    s2 = re.sub(r"_+", "_", s2).strip("_")
    if not s2:
        s2 = "noname"
    if s2[0].isdigit():
        s2 = "m_" + s2
    return prefix + s2


# -----------------------------
# Region helpers (R shell + z interval)
# -----------------------------
def t_interval_inside_z(origin, direction, zmin, zmax) -> Optional[Tuple[float, float]]:
    oz = origin[2]
    dz = direction[2]

    if abs(dz) < 1e-15:
        if zmin <= oz <= zmax:
            return (0.0, float("inf"))
        return None

    t1 = (zmin - oz) / dz
    t2 = (zmax - oz) / dz
    t_in = min(t1, t2)
    t_out = max(t1, t2)

    if t_out < 0:
        return None
    t_in = max(t_in, 0.0)
    return (t_in, t_out)


def t_interval_inside_radial_shell(origin, direction, rmin, rmax) -> Optional[Tuple[float, float]]:
    ox, oy = origin[0], origin[1]
    dx, dy = direction[0], direction[1]

    a = dx * dx + dy * dy
    b = 2.0 * (ox * dx + oy * dy)
    c0 = ox * ox + oy * oy

    if a < 1e-15:
        r0 = math.sqrt(c0)
        if rmin <= r0 <= rmax:
            return (0.0, float("inf"))
        return None

    def roots_for_radius(R) -> Optional[Tuple[float, float]]:
        c = c0 - R * R
        disc = b * b - 4 * a * c
        if disc < 0:
            return None
        s = math.sqrt(disc)
        return ((-b - s) / (2 * a), (-b + s) / (2 * a))

    out_roots = roots_for_radius(rmax)
    if out_roots is None:
        if c0 > rmax * rmax:
            return None
        out_interval = (0.0, float("inf"))
    else:
        t1, t2 = out_roots
        out_interval = (min(t1, t2), max(t1, t2))

    tA, tB = out_interval
    if math.isfinite(tB) and tB < 0:
        return None
    tA = max(tA, 0.0)

    if rmin <= 0:
        return (tA, tB)

    in_roots = roots_for_radius(rmin)
    if in_roots is None:
        if c0 < rmin * rmin:
            return None
        return (tA, tB)

    forbA, forbB = (min(in_roots), max(in_roots))

    seg1 = (tA, min(tB, forbA) if math.isfinite(tB) else forbA)
    seg2 = (max(tA, forbB), tB)

    def valid(seg):
        s, e = seg
        if not math.isfinite(e):
            return True
        return e > s

    if valid(seg1) and (math.isfinite(seg1[1]) and seg1[1] > seg1[0]):
        return seg1
    if valid(seg2) and (not math.isfinite(seg2[1]) or seg2[1] > seg2[0]):
        return seg2
    return None


def t_interval_inside_region(origin, direction, rmin, rmax, zmin, zmax) -> Optional[Tuple[float, float]]:
    z_int = t_interval_inside_z(origin, direction, zmin, zmax)
    r_int = t_interval_inside_radial_shell(origin, direction, rmin, rmax)
    if z_int is None or r_int is None:
        return None

    t0 = max(z_int[0], r_int[0])
    t1 = min(z_int[1], r_int[1])
    if math.isfinite(t1) and t1 <= t0:
        return None
    if t1 <= 0:
        return None
    return (t0, t1)


# -----------------------------
# Ray tracing
# -----------------------------
def trace_ray(
    ROOT,
    geo,
    origin,
    direction,
    rmin_cm: float,
    rmax_cm: float,
    zmin_cm: float,
    zmax_cm: float,
    collect_patterns: List[str],
    expanded_collect_volnames: Set[str],
    max_steps: int = 200000,
    debug_first_steps: int = 0,
    record_segments: bool = False,
) -> Tuple[Dict[str, float], float, List[Tuple], List[Dict[str, Any]]]:
    """
    Collect if either:
      - matches collect_patterns (node/path/volume), OR
      - current volume name is in expanded_collect_volnames (from GDML physvol expansion)

    Returns:
      contrib_dict, total, debug_rows, segments_rows
    """
    nav = geo.GetCurrentNavigator()
    if not nav:
        nav = geo.AddNavigator()

    interval = t_interval_inside_region(origin, direction, rmin_cm, rmax_cm, zmin_cm, zmax_cm)
    if interval is None:
        return {}, 0.0, [], []

    t_in, t_out = interval
    if not math.isfinite(t_out) or t_out <= t_in:
        return {}, 0.0, [], []

    ox, oy, oz = origin
    dx, dy, dz = direction

    sx = ox + dx * t_in
    sy = oy + dy * t_in
    sz = oz + dz * t_in

    nav.SetCurrentPoint(sx, sy, sz)
    nav.SetCurrentDirection(dx, dy, dz)
    nav.FindNode(sx, sy, sz)

    contrib: Dict[str, float] = defaultdict(float)
    total = 0.0

    traveled = t_in
    steps = 0
    eps = 1e-6
    debug_rows: List[Tuple] = []
    segments: List[Dict[str, Any]] = []

    # pre-normalize expanded names for robust match
    expanded_norm = {norm_name(v) for v in expanded_collect_volnames}

    while traveled < t_out and steps < max_steps:
        steps += 1

        node = nav.GetCurrentNode()
        node_name = str(node.GetName()) if node else "None"

        path = ""
        if hasattr(nav, "GetPath"):
            try:
                path = str(nav.GetPath())
            except Exception:
                path = ""

        vol = nav.GetCurrentVolume()
        vol_name = str(vol.GetName()) if vol else "None"
        vol_name_n = norm_name(vol_name)

        by_patterns = is_collected_by_patterns(node_name, path, vol_name, collect_patterns)
        by_expanded = (vol_name in expanded_collect_volnames) or (vol_name_n in expanded_norm)

        do_collect = by_patterns or by_expanded

        mat_name = "Outside"
        x0 = None
        if vol and vol.GetMedium() and vol.GetMedium().GetMaterial():
            mat = vol.GetMedium().GetMaterial()
            mat_name = str(mat.GetName())
            try:
                x0 = float(mat.GetRadLen())  # cm
            except Exception:
                x0 = None
            if x0 is not None and x0 <= 0:
                x0 = None

        remaining = float(t_out - traveled)
        if remaining <= 0:
            break

        nav.FindNextBoundary(remaining)
        step = float(nav.GetStep())

        if debug_first_steps > 0 and len(debug_rows) < debug_first_steps:
            debug_rows.append((step, node_name, vol_name, mat_name, do_collect, by_patterns, by_expanded, path))

        if step <= 0:
            nav.SetStep(eps)
            nav.Step()
            traveled += eps
            continue

        seg_contrib = 0.0
        if do_collect and x0 is not None:
            seg_contrib = step / x0
            contrib[mat_name] += seg_contrib
            total += seg_contrib

        if record_segments:
            t0 = traveled
            t1 = min(traveled + step, t_out)
            x0p, y0p, z0p = (ox + dx * t0, oy + dy * t0, oz + dz * t0)
            x1p, y1p, z1p = (ox + dx * t1, oy + dy * t1, oz + dz * t1)
            r0 = math.sqrt(x0p * x0p + y0p * y0p)
            r1 = math.sqrt(x1p * x1p + y1p * y1p)

            segments.append({
                "ray_id": None,
                "eta": None,
                "phi": None,
                "t_start_cm": t0,
                "t_end_cm": t1,
                "step_cm": step,
                "x_start_cm": x0p,
                "y_start_cm": y0p,
                "z_start_cm": z0p,
                "r_start_cm": r0,
                "x_end_cm": x1p,
                "y_end_cm": y1p,
                "z_end_cm": z1p,
                "r_end_cm": r1,
                "node_name": node_name,
                "node_name_norm": norm_name(node_name),
                "volume_name": vol_name,
                "volume_name_norm": vol_name_n,
                "path": path,
                "material": mat_name,
                "radlen_cm": x0 if x0 is not None else np.nan,
                "collect": bool(do_collect),
                "collect_by_patterns": bool(by_patterns),
                "collect_by_gdml_expansion": bool(by_expanded),
                "x_over_x0": seg_contrib,
            })

        nav.SetStep(step)
        nav.Step()
        traveled += step

        if nav.IsOutside():
            break

    return dict(contrib), float(total), debug_rows, segments


# -----------------------------
# Curved trajectory (helix/arc in solenoidal Bz) for traversal tables
# Approximated by straight chords (polyline).
# -----------------------------
def _norm3(v: Tuple[float, float, float]) -> Tuple[float, float, float]:
    x, y, z = v
    n = math.sqrt(x*x + y*y + z*z)
    if n <= 0:
        return (0.0, 0.0, 0.0)
    return (x/n, y/n, z/n)


def trace_segment_finite(
    ROOT,
    geo,
    start_xyz: Tuple[float, float, float],
    dir_xyz: Tuple[float, float, float],
    length_cm: float,
    collect_patterns: List[str],
    expanded_collect_volnames: Set[str],
    max_steps: int = 200000,
    record_segments: bool = False,
    s_start_cm: float = 0.0,
) -> Tuple[Dict[str, float], float, List[Dict[str, Any]]]:
    """Trace a *finite* straight segment using TGeoNavigator boundary stepping."""
    nav = geo.GetCurrentNavigator()
    if not nav:
        nav = geo.AddNavigator()

    sx, sy, sz = start_xyz
    dx, dy, dz = _norm3(dir_xyz)

    nav.SetCurrentPoint(float(sx), float(sy), float(sz))
    nav.SetCurrentDirection(float(dx), float(dy), float(dz))
    nav.FindNode(float(sx), float(sy), float(sz))

    contrib: Dict[str, float] = defaultdict(float)
    total = 0.0
    traveled = 0.0
    steps = 0
    eps = 1e-6
    segments: List[Dict[str, Any]] = []

    expanded_norm = {norm_name(v) for v in expanded_collect_volnames}

    while traveled < length_cm and steps < max_steps:
        steps += 1

        node = nav.GetCurrentNode()
        node_name = str(node.GetName()) if node else "None"

        path = ""
        if hasattr(nav, "GetPath"):
            try:
                path = str(nav.GetPath())
            except Exception:
                path = ""

        vol = nav.GetCurrentVolume()
        vol_name = str(vol.GetName()) if vol else "None"
        vol_name_n = norm_name(vol_name)

        by_patterns = is_collected_by_patterns(node_name, path, vol_name, collect_patterns)
        by_expanded = (vol_name in expanded_collect_volnames) or (vol_name_n in expanded_norm)
        do_collect = by_patterns or by_expanded

        mat_name = "Outside"
        x0 = None
        if vol and vol.GetMedium() and vol.GetMedium().GetMaterial():
            mat = vol.GetMedium().GetMaterial()
            mat_name = str(mat.GetName())
            try:
                x0 = float(mat.GetRadLen())  # cm
            except Exception:
                x0 = None
            if x0 is not None and x0 <= 0:
                x0 = None

        remaining = float(length_cm - traveled)
        if remaining <= 0:
            break

        nav.FindNextBoundary(remaining)
        step = float(nav.GetStep())

        if step <= 0:
            nav.SetStep(eps)
            nav.Step()
            traveled += eps
            continue

        seg_contrib = 0.0
        if do_collect and x0 is not None:
            seg_contrib = step / x0
            contrib[mat_name] += seg_contrib
            total += seg_contrib

        if record_segments:
            s0 = s_start_cm + traveled
            s1 = s_start_cm + min(traveled + step, length_cm)

            t0 = traveled
            t1 = min(traveled + step, length_cm)

            x0p, y0p, z0p = (sx + dx * t0, sy + dy * t0, sz + dz * t0)
            x1p, y1p, z1p = (sx + dx * t1, sy + dy * t1, sz + dz * t1)
            r0 = math.sqrt(x0p * x0p + y0p * y0p)
            r1 = math.sqrt(x1p * x1p + y1p * y1p)

            segments.append({
                "ray_id": None,
                "eta": None,
                "phi": None,
                "t_start_cm": s0,
                "t_end_cm": s1,
                "step_cm": step,
                "x_start_cm": x0p,
                "y_start_cm": y0p,
                "z_start_cm": z0p,
                "r_start_cm": r0,
                "x_end_cm": x1p,
                "y_end_cm": y1p,
                "z_end_cm": z1p,
                "r_end_cm": r1,
                "node_name": node_name,
                "node_name_norm": norm_name(node_name),
                "volume_name": vol_name,
                "volume_name_norm": vol_name_n,
                "path": path,
                "material": mat_name,
                "radlen_cm": x0 if x0 is not None else np.nan,
                "collect": bool(do_collect),
                "collect_by_patterns": bool(by_patterns),
                "collect_by_gdml_expansion": bool(by_expanded),
                "x_over_x0": seg_contrib,
            })

        nav.SetStep(step)
        nav.Step()
        traveled += step

        if nav.IsOutside():
            break

    return dict(contrib), float(total), segments


def trace_arc_helix_for_table(
    ROOT,
    geo,
    origin: Tuple[float, float, float],
    eta: float,
    phi: float,
    rmin_cm: float,
    rmax_cm: float,
    zmin_cm: float,
    zmax_cm: float,
    collect_patterns: List[str],
    expanded_collect_volnames: Set[str],
    *,
    pt_gev: float = 1.0,
    bz_tesla: float = 2.0,
    charge: int = 1,
    arc_step_mm: float = 1.0,
    max_chords: int = 200000,
    record_segments: bool = False,
) -> Tuple[Dict[str, float], float, List[Dict[str, Any]]]:
    """Trace material along an approximate helix by polyline chords."""
    if abs(bz_tesla) < 1e-15 or abs(pt_gev) < 1e-15 or abs(charge) < 1e-15:
        direction = eta_phi_to_dir(float(eta), float(phi))
        c, t, _dbg, segs = trace_ray(
            ROOT, geo, origin, direction,
            rmin_cm=float(rmin_cm), rmax_cm=float(rmax_cm),
            zmin_cm=float(zmin_cm), zmax_cm=float(zmax_cm),
            collect_patterns=collect_patterns,
            expanded_collect_volnames=expanded_collect_volnames,
            max_steps=int(max_chords),
            debug_first_steps=0,
            record_segments=record_segments,
        )
        return c, t, segs

    R_m = float(pt_gev) / (0.299792458 * abs(float(charge)) * abs(float(bz_tesla)))
    R_cm = R_m * 100.0

    sgn = 1.0 if (float(charge) * float(bz_tesla)) > 0 else -1.0

    ox, oy, oz = origin
    cx = ox + sgn * R_cm * math.sin(phi)
    cy = oy - sgn * R_cm * math.cos(phi)
    alpha = math.atan2(oy - cy, ox - cx)

    dsT_cm = float(arc_step_mm) / 10.0
    if dsT_cm <= 0:
        dsT_cm = 0.1

    dphiT = dsT_cm / R_cm
    dalpha = -sgn * dphiT
    dz_step = (R_cm * dalpha * math.sinh(float(eta)))

    def in_bounds(xx, yy, zz):
        rr = math.sqrt(xx*xx + yy*yy)
        return (rmin_cm <= rr <= rmax_cm) and (zmin_cm <= zz <= zmax_cm)

    x, y, z = ox, oy, oz
    s_cm = 0.0

    contrib_all: Dict[str, float] = defaultdict(float)
    total_all = 0.0
    seg_rows: List[Dict[str, Any]] = []

    for _ in range(int(max_chords)):
        alpha2 = alpha + dalpha
        x2 = cx + R_cm * math.cos(alpha2)
        y2 = cy + R_cm * math.sin(alpha2)
        z2 = z + dz_step

        dx, dy, dz = (x2 - x, y2 - y, z2 - z)
        L = math.sqrt(dx*dx + dy*dy + dz*dz)
        if L <= 0:
            break

        rr2 = math.sqrt(x2*x2 + y2*y2)
        out = (rr2 > rmax_cm) or (z2 < zmin_cm) or (z2 > zmax_cm)

        if out:
            if in_bounds(x, y, z) or in_bounds(x2, y2, z2):
                c, t, segs = trace_segment_finite(
                    ROOT, geo, (x, y, z), (dx, dy, dz), L,
                    collect_patterns, expanded_collect_volnames,
                    max_steps=int(max_chords),
                    record_segments=record_segments,
                    s_start_cm=s_cm,
                )
                for k, v in c.items():
                    contrib_all[k] += v
                total_all += t
                if record_segments:
                    seg_rows.extend(segs)
            break

        c, t, segs = trace_segment_finite(
            ROOT, geo, (x, y, z), (dx, dy, dz), L,
            collect_patterns, expanded_collect_volnames,
            max_steps=int(max_chords),
            record_segments=record_segments,
            s_start_cm=s_cm,
        )
        for k, v in c.items():
            contrib_all[k] += v
        total_all += t
        if record_segments:
            seg_rows.extend(segs)

        x, y, z = x2, y2, z2
        alpha = alpha2
        s_cm += L

    return dict(contrib_all), float(total_all), seg_rows



# -----------------------------
# ROOT output helpers
# -----------------------------
def write_root_outputs(
    ROOT,
    root_out_path: str,
    x_min: float,
    x_max: float,
    x_label: str,
    per_mat_1d: Dict[str, np.ndarray],
    stacked_mats_ordered: List[str],
    title_1d: str,
    meta_dict: Dict[str, str],
    th2_total: Optional[np.ndarray] = None,
    th2_eta_edges: Optional[Tuple[float, float, int]] = None,
    th2_phi_edges: Optional[Tuple[float, float, int]] = None,
):
    n = len(next(iter(per_mat_1d.values()))) if per_mat_1d else 0
    tf = ROOT.TFile(root_out_path, "RECREATE")
    tf.cd()

    h_total = ROOT.TH1D("h_total", title_1d, int(n), float(x_min), float(x_max))
    h_total.GetXaxis().SetTitle(x_label)
    h_total.GetYaxis().SetTitle("X/X0")
    h_total.Sumw2()

    total_arr = np.zeros(int(n), dtype=float)
    for m in stacked_mats_ordered:
        total_arr += per_mat_1d[m]
    for i in range(int(n)):
        h_total.SetBinContent(i + 1, float(total_arr[i]))
    h_total.Write()

    for m in stacked_mats_ordered:
        h = ROOT.TH1D(sanitize_root_name(m), f"{m};{x_label};X/X0", int(n), float(x_min), float(x_max))
        h.Sumw2()
        arr = per_mat_1d[m]
        for i in range(int(n)):
            h.SetBinContent(i + 1, float(arr[i]))
        h.Write()

    if th2_total is not None and th2_eta_edges and th2_phi_edges:
        eta_min, eta_max, n_eta = th2_eta_edges
        phi_min, phi_max, n_phi = th2_phi_edges
        h2 = ROOT.TH2D(
            "h2_total_eta_phi",
            "Total X/X0;eta;phi [rad]",
            int(n_eta), float(eta_min), float(eta_max),
            int(n_phi), float(phi_min), float(phi_max),
        )
        for i in range(int(n_eta)):
            for j in range(int(n_phi)):
                h2.SetBinContent(i + 1, j + 1, float(th2_total[i, j]))
        h2.Write()

    for k, v in (meta_dict or {}).items():
        ROOT.TNamed(str(k), str(v)).Write()

    tf.Close()


# -----------------------------
# Traversal table outputs
# -----------------------------
def _postprocess_segments_df(
    df,
    threshold: float = 0.0001,
    gap_mm: float = 1.0,
    barrel_name: str = "barrel_1",
):
    """
    Input: DataFrame containing at least columns:
      - ray_id
      - x_over_x0
      - t_start_cm
      - t_end_cm
      - node_name
      - r_start_cm (for export)

    Produces:
      df_filtered: rows with x_over_x0 >= threshold
      df_merged: within each ray_id, merges by rule:
          if (next.t_start_cm - curr.t_end_cm) < gap_cm OR (next.node_name == barrel_name)
          then delete next row and add next.x_over_x0 to curr.x_over_x0
      df_export: columns: node_name (quoted), ",", r_start_cm, ",", x_over_x0
    """
    import pandas as pd  # type: ignore

    df0 = df.copy()

    required = ["eta", "ray_id", "x_over_x0", "t_start_cm", "t_end_cm", "node_name", "r_start_cm"]
    missing = [c for c in required if c not in df0.columns]
    if missing:
        raise ValueError(f"Segments table missing required columns: {missing}")

    for c in ["x_over_x0", "t_start_cm", "t_end_cm", "r_start_cm"]:
        df0[c] = pd.to_numeric(df0[c], errors="coerce")

    df_filtered = df0.loc[(df0["x_over_x0"].notna()) & (df0["x_over_x0"] >= float(threshold))].copy()

    gap_cm = float(gap_mm) / 10.0  # mm -> cm

    def merge_one_ray(g):
        g = g.reset_index(drop=True).copy()
        i = 0
        while i < len(g) - 1:
            curr_t_end = g.at[i, "t_end_cm"]
            next_t_start = g.at[i + 1, "t_start_cm"]
            next_node = g.at[i + 1, "node_name"]

            is_barrel = (str(next_node).strip() == barrel_name) if pd.notna(next_node) else False

            close_enough = False
            if pd.notna(curr_t_end) and pd.notna(next_t_start):
                close_enough = (float(next_t_start) - float(curr_t_end)) < gap_cm

            if close_enough or is_barrel:
                curr_x = g.at[i, "x_over_x0"]
                next_x = g.at[i + 1, "x_over_x0"]
                curr_x = float(curr_x) if pd.notna(curr_x) else 0.0
                next_x = float(next_x) if pd.notna(next_x) else 0.0
                g.at[i, "x_over_x0"] = curr_x + next_x

                # If this merge happened because the next layer is within the gap
                # and the next node_name contains "Sensor", keep the Sensor name.
                if close_enough and pd.notna(next_node) and ("Sensor" in str(next_node)):
                    g.at[i, "node_name"] = next_node

                g = g.drop(index=i + 1).reset_index(drop=True)
                continue

            i += 1
        return g

    df_merged = (
        df_filtered
        .groupby(["eta", "ray_id"], sort=False, dropna=False, group_keys=False)
        .apply(merge_one_ray)
        .reset_index(drop=True)
    )

    def quote(v):
        if pd.isna(v):
            return '""'
        return f"\"{v}\""

    df_export = pd.DataFrame({
    "eta": df_merged["eta"],
    "node_name": df_merged["node_name"].map(quote),
    ",_1": [","] * len(df_merged),
    "r_start_cm": df_merged["r_start_cm"],
    ",_2": [","] * len(df_merged),
    "x_over_x0": df_merged["x_over_x0"],
})

    return df_filtered, df_merged, df_export


def save_traversal_tables_postprocessed(
    segments_rows: List[Dict[str, Any]],
    out_prefix: str,
    threshold: float = 0.0001,
    gap_mm: float = 1.0,
    barrel_name: str = "barrel_1",
):
    """
    Writes a single XLSX:
      - out_prefix.xlsx with sheets:
          * segments (raw)
          * by_material, by_volume (if available)
          * segments_filtered
          * segments_merged
          * segments_merged_export

    No intermediate XLSX or TXT is produced.
    Post-processing (filtered/merged/export) is done independently for each eta.
    """
    xlsx_path = out_prefix + ".xlsx"

    import pandas as pd  # type: ignore

    if not segments_rows:
        # Create empty xlsx with expected sheets
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            pd.DataFrame([]).to_excel(writer, index=False, sheet_name="segments")
            pd.DataFrame([]).to_excel(writer, index=False, sheet_name="segments_filtered")
            pd.DataFrame([]).to_excel(writer, index=False, sheet_name="segments_merged")
            pd.DataFrame([]).to_excel(writer, index=False, sheet_name="segments_merged_export")
        return

    df = pd.DataFrame(segments_rows)

    # Legacy-style summaries (unchanged)
    summaries = {}
    if "ray_id" in df.columns and "x_over_x0" in df.columns:
        if all(c in df.columns for c in ["eta", "phi", "material"]):
            summaries["by_material"] = (
                df.groupby(["ray_id", "eta", "phi", "material"], dropna=False)["x_over_x0"]
                  .sum().reset_index()
                  .sort_values(["ray_id", "x_over_x0"], ascending=[True, False])
            )
        if all(c in df.columns for c in ["eta", "phi", "volume_name_norm"]):
            summaries["by_volume"] = (
                df.groupby(["ray_id", "eta", "phi", "volume_name_norm"], dropna=False)["x_over_x0"]
                  .sum().reset_index()
                  .sort_values(["ray_id", "x_over_x0"], ascending=[True, False])
            )

    df_filtered, df_merged, df_export = _postprocess_segments_df(
        df, threshold=threshold, gap_mm=gap_mm, barrel_name=barrel_name
    )

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="segments")
        for name, sdf in summaries.items():
            sdf.to_excel(writer, index=False, sheet_name=name[:31])

        df_filtered.to_excel(writer, index=False, sheet_name="segments_filtered")
        df_merged.to_excel(writer, index=False, sheet_name="segments_merged")
        df_export.to_excel(writer, index=False, sheet_name="segments_merged_export")

    # Patch comma headers to be exactly ","
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(xlsx_path)
        ws = wb["segments_merged_export"]
        # headers are row 1: eta, node_name, ,_1, r_start_cm, ,_2, x_over_x0
        if ws.cell(1, 3).value in (",_1", ","):
            ws.cell(1, 3).value = ","
        if ws.cell(1, 5).value in (",_2", ","):
            ws.cell(1, 5).value = ","
        wb.save(xlsx_path)
    except Exception:
        pass

    # ---------------------------------------------------------
    # Save export sheet also as txt
    # ---------------------------------------------------------
    txt_export_path = out_prefix + "_export.txt"

    try:
        with open(txt_export_path, "w", encoding="utf-8") as f:

            # header
            f.write("eta,node_name,r_start_cm,x_over_x0\n")

            for _, row in df_export.iterrows():
                eta = row["eta"]
                node = row["node_name"]   # уже в кавычках
                r = row["r_start_cm"]
                x = row["x_over_x0"]

                f.write(f"{eta},{node},{r},{x}\n")

        print(f"[INFO] export txt written: {txt_export_path}")

    except Exception as e:
        print(f"[WARN] failed writing export txt: {e}")

# -----------------------------
# Main


# -----------------------------
def main():
    ap = argparse.ArgumentParser()

    ap.add_argument("--gdml", required=True, help="Path to GDML file.")

    ap.add_argument("--plot-vs", choices=["eta", "phi"], default="eta")

    ap.add_argument("--eta-min", type=float, default=-4.0)
    ap.add_argument("--eta-max", type=float, default=+4.0)
    ap.add_argument("--n-eta", type=int, default=161)

    ap.add_argument("--phi-min", type=float, default=0.0)
    ap.add_argument("--phi-max", type=float, default=0.0)
    ap.add_argument("--n-phi", type=int, default=1)

    ap.add_argument("--origin-x", type=float, default=0.0)
    ap.add_argument("--origin-y", type=float, default=30.0)
    ap.add_argument("--origin-z", type=float, default=0.0)

    ap.add_argument("--rmin", type=float, default=0.0)
    ap.add_argument("--rmax", type=float, default=90.0)
    ap.add_argument("--zmin", type=float, default=-300.0)
    ap.add_argument("--zmax", type=float, default=300.0)

    # Old pattern-based collector (still supported)
    ap.add_argument("--collect-nodes", action="append", default=None,
                    help="Pattern(s) for node/path/volume matching. Repeatable. Supports wildcards.")
    ap.add_argument("--collect-nodes-csv", default=None)

    # New: physvol expansion from GDML (this is what fixes FT3V_2 assemblies)
    ap.add_argument("--collect-physvol", action="append", default=None,
                    help="Physvol name(s) to collect inside (and all descendants), expanded from GDML structure.")
    ap.add_argument("--collect-physvols-csv", default=None)

    ap.add_argument("--max-steps", type=int, default=200000)
    ap.add_argument("--top-materials", type=int, default=15)

    ap.add_argument("--out", default="matbud.png")
    ap.add_argument("--root-out", default="matbud.root")

    ap.add_argument("--make-th2", action="store_true")
    ap.add_argument("--th2-eta-min", type=float, default=None)
    ap.add_argument("--th2-eta-max", type=float, default=None)
    ap.add_argument("--th2-n-eta", type=int, default=None)
    ap.add_argument("--th2-phi-min", type=float, default=None)
    ap.add_argument("--th2-phi-max", type=float, default=None)
    ap.add_argument("--th2-n-phi", type=int, default=None)

    ap.add_argument("--table-mode", choices=["pairwise", "cartesian"], default="pairwise")
    ap.add_argument("--table-etas", default=None)
    ap.add_argument("--table-phis", default=None)
    ap.add_argument("--table-out-prefix", default="ray_traversal")

    # Post-processing of traversal xlsx (segments_filtered/merged/export)
    ap.add_argument("--segments-threshold", type=float, default=0.0001,
                help="Filter threshold for segments: keep rows with x_over_x0 >= threshold.")
    ap.add_argument("--merge-gap-mm", type=float, default=1.0,
                help="Merge if next.t_start_cm - curr.t_end_cm < gap (in mm). Default: 1 mm.")
    ap.add_argument("--merge-barrel-name", default="barrel_1",
                help="If next node_name equals this, merge it into previous row (always). Default: barrel_1.")


    # Curved-track trajectory for traversal tables (polyline chords in solenoidal Bz)
    ap.add_argument("--pt", type=float, default=1.0,
                    help="pT in GeV/c for curved trajectory in table scan (default: 1.0).")
    ap.add_argument("--bz", type=float, default=2.0,
                    help="Solenoidal Bz in Tesla for curved trajectory in table scan (default: 2.0).")
    ap.add_argument("--charge", type=int, default=1,
                    help="Particle charge (+1 or -1) for curved trajectory in table scan (default: +1).")
    ap.add_argument("--arc-step-mm", type=float, default=1.0,
                    help="Chord step size along transverse arc in mm (default: 1.0). Smaller is more accurate/slower.")
    ap.add_argument("--debug-ray", action="store_true")
    ap.add_argument("--debug-eta", type=float, default=0.0)
    ap.add_argument("--debug-phi", type=float, default=0.0)
    ap.add_argument("--debug-steps", type=int, default=120)

    args = ap.parse_args()

    if args.rmax <= args.rmin:
        raise SystemExit(f"Invalid radii: rmax ({args.rmax}) must be > rmin ({args.rmin}).")
    if args.zmax <= args.zmin:
        raise SystemExit(f"Invalid z-range: zmax ({args.zmax}) must be > zmin ({args.zmin}).")

    default_collect_physvols = ["A3IP_1", "TRKV_2", "IOTOFV_2", "FT3V_2"]
    default_collect_patterns = []  # keep empty by default; physvol expansion is primary now

    # collect patterns (optional)
    collect_patterns: List[str] = []
    if args.collect_nodes:
        collect_patterns.extend(args.collect_nodes)
    if args.collect_nodes_csv:
        collect_patterns.extend(parse_csv_strings(args.collect_nodes_csv))
    if not collect_patterns:
        collect_patterns = default_collect_patterns

    # collect physvols (expanded)
    collect_physvols: List[str] = []
    if args.collect_physvol:
        collect_physvols.extend(args.collect_physvol)
    if args.collect_physvols_csv:
        collect_physvols.extend(parse_csv_strings(args.collect_physvols_csv))
    if not collect_physvols:
        collect_physvols = default_collect_physvols

    # Parse GDML structure for physvol expansion
    physvol_to_ref, container_to_children, _all_containers = parse_gdml_structure(args.gdml)
    expanded_collect_volnames: Set[str] = set()

    missing_physvols = []
    for pv in collect_physvols:
        ref = physvol_to_ref.get(pv)
        if not ref:
            missing_physvols.append(pv)
            continue
        expanded_collect_volnames |= expand_descendants(ref, container_to_children)

    if missing_physvols:
        print(f"[WARN] These physvol names were not found in GDML <structure>: {missing_physvols}")
        print("       (If you rely on physvol expansion, check the exact names in GDML.)")

    # Quick info: helpful when debugging FT3V_2
    if "FT3V_2" in collect_physvols:
        # show that we expanded to boxTile...
        interesting = [n for n in expanded_collect_volnames if n in ("boxTileAir", "boxTileChip", "boxTileSensor")]
        if interesting:
            print(f"[INFO] FT3V_2 expansion includes: {sorted(interesting)}")
        else:
            # still okay: maybe names differ slightly; user asked these exact ones
            pass

    ROOT, geo = load_root_geometry(args.gdml)
    origin = (float(args.origin_x), float(args.origin_y), float(args.origin_z))

    # Debug one ray
    if args.debug_ray:
        direction = eta_phi_to_dir(float(args.debug_eta), float(args.debug_phi))
        contrib, total, dbg, _segs = trace_ray(
            ROOT, geo, origin, direction,
            rmin_cm=float(args.rmin), rmax_cm=float(args.rmax),
            zmin_cm=float(args.zmin), zmax_cm=float(args.zmax),
            collect_patterns=collect_patterns,
            expanded_collect_volnames=expanded_collect_volnames,
            max_steps=int(args.max_steps),
            debug_first_steps=int(args.debug_steps),
            record_segments=False,
        )
        print(f"[DEBUG] origin={origin} cm, eta={args.debug_eta}, phi={args.debug_phi} rad")
        print(f"[DEBUG] region: R in [{args.rmin},{args.rmax}] cm, z in [{args.zmin},{args.zmax}] cm")
        print(f"[DEBUG] collect physvols: {collect_physvols}")
        print(f"[DEBUG] expanded volume/assembly count: {len(expanded_collect_volnames)}")
        if collect_patterns:
            print(f"[DEBUG] collect patterns: {collect_patterns}")
        for i, row in enumerate(dbg):
            step, node_name, vol_name, mat_name, do_collect, by_pat, by_exp, path = row
            print(f"  step {i:03d}: step={step:.6g} cm | collect={do_collect} "
                  f"(patterns={by_pat}, gdml_exp={by_exp}) | "
                  f"node={node_name} (norm={norm_name(node_name)}) | "
                  f"vol={vol_name} (norm={norm_name(vol_name)}) | mat={mat_name} | path={path}")
        print(f"[DEBUG] collected total X/X0 = {total:.6g}")
        if contrib:
            top_dbg = sorted(contrib.items(), key=lambda kv: kv[1], reverse=True)[:20]
            print("[DEBUG] top materials:")
            for m, v in top_dbg:
                print(f"   {m:30s}  {v:.6g}")
        return

    # Traversal tables
    table_etas = parse_csv_floats(args.table_etas)
    table_phis = parse_csv_floats(args.table_phis)
    if table_etas or table_phis:
        if not (table_etas and table_phis):
            print("[WARN] To write traversal tables you must provide BOTH --table-etas and --table-phis.")
        else:
            rays: List[Tuple[float, float]] = []
            if args.table_mode == "pairwise":
                if len(table_etas) != len(table_phis):
                    raise SystemExit(
                        f"--table-mode pairwise requires same lengths: "
                        f"len(table-etas)={len(table_etas)} != len(table-phis)={len(table_phis)}"
                    )
                rays = list(zip(table_etas, table_phis))
            else:
                for e in table_etas:
                    for p in table_phis:
                        rays.append((e, p))

            all_rows: List[Dict[str, Any]] = []
            for rid, (eta, phi) in enumerate(rays, start=1):
                direction = eta_phi_to_dir(float(eta), float(phi))

                _c, _t, segs = trace_arc_helix_for_table(
                    ROOT, geo, origin,
                    float(eta), float(phi),
                    rmin_cm=float(args.rmin), rmax_cm=float(args.rmax),
                    zmin_cm=float(args.zmin), zmax_cm=float(args.zmax),
                    collect_patterns=collect_patterns,
                    expanded_collect_volnames=expanded_collect_volnames,
                    pt_gev=float(args.pt),
                    bz_tesla=float(args.bz),
                    charge=int(args.charge),
                    arc_step_mm=float(args.arc_step_mm),
                    max_chords=int(args.max_steps),
                    record_segments=True,
                )
                for row in segs:
                    row["ray_id"] = rid
                    row["eta"] = float(eta)
                    row["phi"] = float(phi)
                    all_rows.append(row)

            save_traversal_tables_postprocessed(all_rows, args.table_out_prefix,
                                             threshold=float(args.segments_threshold),
                                             gap_mm=float(args.merge_gap_mm),
                                             barrel_name=str(args.merge_barrel_name))
            print(f"[OK] Saved traversal tables: {args.table_out_prefix}.xlsx and {args.table_out_prefix}_export.txt")

    # 1D scan grids
    etas = linspace_including_end(args.eta_min, args.eta_max, args.n_eta)
    phis = linspace_including_end(args.phi_min, args.phi_max, args.n_phi)

    mat_integral = Counter()
    per_mat_x: Dict[str, np.ndarray] = {}

    if args.plot_vs == "eta":
        x_vals = etas
        x_label_root = "eta"
        xlab_plot = r"$\eta$"
        x_min, x_max = float(args.eta_min), float(args.eta_max)

        per_mat_eta = defaultdict(lambda: np.zeros_like(etas, dtype=float))
        for i, eta in enumerate(etas):
            accum = defaultdict(float)
            for phi in phis:
                direction = eta_phi_to_dir(float(eta), float(phi))
                contrib, _total, _dbg, _ = trace_ray(
                    ROOT, geo, origin, direction,
                    rmin_cm=float(args.rmin), rmax_cm=float(args.rmax),
                    zmin_cm=float(args.zmin), zmax_cm=float(args.zmax),
                    collect_patterns=collect_patterns,
                    expanded_collect_volnames=expanded_collect_volnames,
                    max_steps=int(args.max_steps),
                    record_segments=False,
                )
                for m, v in contrib.items():
                    accum[m] += v
            for m, v in accum.items():
                vmean = v / len(phis)
                per_mat_eta[m][i] = vmean
                mat_integral[m] += vmean

        per_mat_x = dict(per_mat_eta)
        title_1d = (
            rf"X/X0 vs $\eta$ | origin=({origin[0]:.1f},{origin[1]:.1f},{origin[2]:.1f}) cm "
            rf"| R∈[{args.rmin:.0f},{args.rmax:.0f}] cm, z∈[{args.zmin:.0f},{args.zmax:.0f}] cm "
            rf"| $\phi\in[{args.phi_min:.3g},{args.phi_max:.3g}]$ (n={len(phis)}) "
            rf"| vols={collect_physvols}"
        )
    else:
        x_vals = phis
        x_label_root = "phi [rad]"
        xlab_plot = r"$\phi$ [rad]"
        x_min, x_max = float(args.phi_min), float(args.phi_max)

        per_mat_phi = defaultdict(lambda: np.zeros_like(phis, dtype=float))
        for i, phi in enumerate(phis):
            accum = defaultdict(float)
            for eta in etas:
                direction = eta_phi_to_dir(float(eta), float(phi))
                contrib, _total, _dbg, _ = trace_ray(
                    ROOT, geo, origin, direction,
                    rmin_cm=float(args.rmin), rmax_cm=float(args.rmax),
                    zmin_cm=float(args.zmin), zmax_cm=float(args.zmax),
                    collect_patterns=collect_patterns,
                    expanded_collect_volnames=expanded_collect_volnames,
                    max_steps=int(args.max_steps),
                    record_segments=False,
                )
                for m, v in contrib.items():
                    accum[m] += v
            for m, v in accum.items():
                vmean = v / len(etas)
                per_mat_phi[m][i] = vmean
                mat_integral[m] += vmean

        per_mat_x = dict(per_mat_phi)
        title_1d = (
            rf"X/X0 vs $\phi$ | origin=({origin[0]:.1f},{origin[1]:.1f},{origin[2]:.1f}) cm "
            rf"| R∈[{args.rmin:.0f},{args.rmax:.0f}] cm, z∈[{args.zmin:.0f},{args.zmax:.0f}] cm "
            rf"| $\eta\in[{args.eta_min:.3g},{args.eta_max:.3g}]$ (n={len(etas)}) "
            rf"| vols={collect_physvols}"
        )

    top = [m for m, _ in mat_integral.most_common(int(args.top_materials))]
    stacked_mats: List[str] = []
    stacked_vals: List[np.ndarray] = []
    other = np.zeros_like(x_vals, dtype=float)

    for m, arr in per_mat_x.items():
        if m in top:
            stacked_mats.append(m)
            stacked_vals.append(arr)
        else:
            other += arr
    if np.any(other > 0):
        stacked_mats.append("Other")
        stacked_vals.append(other)
        per_mat_x["Other"] = other

    if len(stacked_vals) == 0:
        print("[WARN] No material collected for 1D plot (empty stack). Check physvol names / region.")
        plt.figure(figsize=(11, 6))
        plt.plot(x_vals, np.zeros_like(x_vals), label="(no collected material)")
        plt.xlabel(xlab_plot)
        plt.ylabel(r"Material budget $\sum \ell/X_0$")
        plt.title("X/X0 — nothing collected")
        plt.grid(True, which="both", linewidth=0.5, alpha=0.5)
        plt.legend()
        plt.tight_layout()
        plt.savefig(args.out, dpi=200)
        print(f"[OK] Saved PNG: {args.out}")

        meta = {
            "gdml": args.gdml,
            "origin_cm": str(origin),
            "rmin_cm": str(args.rmin), "rmax_cm": str(args.rmax),
            "zmin_cm": str(args.zmin), "zmax_cm": str(args.zmax),
            "collect_physvols": str(collect_physvols),
            "expanded_volume_count": str(len(expanded_collect_volnames)),
            "collect_patterns": str(collect_patterns),
        }
        write_root_outputs(
            ROOT, args.root_out,
            x_min=float(x_min), x_max=float(x_max), x_label=x_label_root,
            per_mat_1d={"(none)": np.zeros_like(x_vals)},
            stacked_mats_ordered=["(none)"],
            title_1d=title_1d,
            meta_dict=meta,
        )
        print(f"[OK] Saved ROOT: {args.root_out}")
        return

    order = np.argsort([float(np.sum(a)) for a in stacked_vals])
    stacked_mats = [stacked_mats[k] for k in order]
    stacked_vals = [stacked_vals[k] for k in order]

    plt.figure(figsize=(11, 6))
    plt.stackplot(x_vals, stacked_vals, labels=stacked_mats, alpha=0.9)
    plt.xlabel(xlab_plot)
    plt.ylabel(r"Material budget $\sum \ell/X_0$")
    plt.title(title_1d)
    plt.grid(True, which="both", linewidth=0.5, alpha=0.5)
    plt.legend(loc="upper right", fontsize=8, ncol=2, frameon=True)
    plt.tight_layout()
    plt.savefig(args.out, dpi=200)
    plt.savefig(args.out+".pdf", dpi=200)
    print(f"[OK] Saved PNG: {args.out}")

    th2_total = None
    th2_eta_edges = None
    th2_phi_edges = None

    if args.make_th2:
        eta_min2 = args.th2_eta_min if args.th2_eta_min is not None else float(args.eta_min)
        eta_max2 = args.th2_eta_max if args.th2_eta_max is not None else float(args.eta_max)
        n_eta2 = args.th2_n_eta if args.th2_n_eta is not None else int(args.n_eta)

        phi_min2 = args.th2_phi_min if args.th2_phi_min is not None else float(args.phi_min)
        phi_max2 = args.th2_phi_max if args.th2_phi_max is not None else float(args.phi_max)
        n_phi2 = args.th2_n_phi if args.th2_n_phi is not None else int(args.n_phi)

        eta_grid = linspace_including_end(eta_min2, eta_max2, n_eta2)
        phi_grid = linspace_including_end(phi_min2, phi_max2, n_phi2)

        th2_total = np.zeros((n_eta2, n_phi2), dtype=float)
        for i, eta in enumerate(eta_grid):
            for j, phi in enumerate(phi_grid):
                direction = eta_phi_to_dir(float(eta), float(phi))
                _c, tot, _dbg, _ = trace_ray(
                    ROOT, geo, origin, direction,
                    rmin_cm=float(args.rmin), rmax_cm=float(args.rmax),
                    zmin_cm=float(args.zmin), zmax_cm=float(args.zmax),
                    collect_patterns=collect_patterns,
                    expanded_collect_volnames=expanded_collect_volnames,
                    max_steps=int(args.max_steps),
                    record_segments=False,
                )
                th2_total[i, j] = tot

        th2_eta_edges = (float(eta_min2), float(eta_max2), int(n_eta2))
        th2_phi_edges = (float(phi_min2), float(phi_max2), int(n_phi2))
        print("[OK] Computed TH2 total X/X0(eta,phi)")

    meta = {
        "gdml": args.gdml,
        "origin_cm": str(origin),
        "rmin_cm": str(args.rmin), "rmax_cm": str(args.rmax),
        "zmin_cm": str(args.zmin), "zmax_cm": str(args.zmax),
        "collect_physvols": str(collect_physvols),
        "expanded_volume_count": str(len(expanded_collect_volnames)),
        "collect_patterns": str(collect_patterns),
        "eta_range_1d": f"{args.eta_min},{args.eta_max},n={args.n_eta}",
        "phi_range_1d": f"{args.phi_min},{args.phi_max},n={args.n_phi}",
        "top_materials": str(args.top_materials),
        "make_th2": str(bool(args.make_th2)),
        "table_mode": str(args.table_mode),
        "table_etas": str(args.table_etas),
        "table_phis": str(args.table_phis),
    }

    per_mat_ordered = {m: per_mat_x[m] for m in stacked_mats if m in per_mat_x}

    write_root_outputs(
        ROOT, args.root_out,
        x_min=float(x_min), x_max=float(x_max), x_label=x_label_root,
        per_mat_1d=per_mat_ordered,
        stacked_mats_ordered=stacked_mats,
        title_1d=title_1d,
        meta_dict=meta,
        th2_total=th2_total,
        th2_eta_edges=th2_eta_edges,
        th2_phi_edges=th2_phi_edges,
    )
    print(f"[OK] Saved ROOT: {args.root_out}")


if __name__ == "__main__":
    main()
